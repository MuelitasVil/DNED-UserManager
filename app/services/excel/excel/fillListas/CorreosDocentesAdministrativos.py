from app.services.excel.utils import getCantOfColumns
from app.services.excel.excel.archvivosExcel import ArchivosExcel
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
import openpyxl 
import os
import csv

'''
La siguiente clase tendra la funcionalidad de recorrer el excel en el cual se
encuentran los datos de los inquilinos.

Ojo : Es necesario que entienda la estructura del excel para entender las
siguientes funciones.
'''

class CorreosDocentesAdministrativos:
    def __init__(self, file = None, files = []):
        # Lectura del excel 
        if file != None:
            self.excel = openpyxl.load_workbook(file)
            print("CREACION CLASE DE CREACION DE LISTADOS DOCENTES Y ADMNISTRATIVOS")
            print(self.excel.sheetnames)
            self.hojas = self.excel.get_sheet_names()

        self.folder_pathDocentes = "docentes"
        if not os.path.exists(self.folder_pathDocentes):
            os.makedirs(self.folder_pathDocentes)

        self.folder_pathAdministrativos = "admistrativos"
        if not os.path.exists(self.folder_pathAdministrativos):
            os.makedirs(self.folder_pathAdministrativos)

        self.UserFiles = files
        self.columnaInicial = 1
        self.filaInicial = 1   

        self.bogota = "SEDE BOGOTÁ"
        self.amazona = "SEDE AMAZONÍA"
        self.caribe = "SEDE CARIBE"
        self.paz = "SEDE DE LA PAZ" # 
        self.manizales = "SEDE MANIZALES" #
        self.medellin = "SEDE MEDELLÍN" # 
        self.orinoquia = "SEDE ORINOQUÍA" #
        self.palmira = "SEDE PALMIRA" #
        self.tumaco = "SEDE TUMACO"

    def FilterDocentesAdministrativos(self):
        
        '''
        Nota : Esta solucion se podria hacer con un arbol binario.
        Sin embargo lo implemente con diccionarios por facilidad

        Estructura : 
        Sedes -> facultades -> unidadres -> profesores

        Dict-Profesores:
        dict = {
            "sede" : {
                "Unidad" : {
                    
                    }
                }
            }

        Estructura : 
        Sedes  -> administrativos

        Dict-Administrativos:
        dict = {
            "sede" : {
                }
            }


        '''

        # Datos : 

        ArchivoDocentesAdministrativos = ArchivosExcel.Docentes 
        information = self.excel["Hoja1"]
        cantOfRows = len(list(information.rows))
        filaInicial = self.filaInicial
 
        #woorkbookOfplans.create_sheet("Informacion General")
        
        # Obtener informacion : 
        dict_Of_Docentes = {}
        dict_Of_Administrativos = {}

        print("OBTENIENDO INFORMACION : ")
        for row in range(filaInicial, cantOfRows):
            
            columnNombreVinculacion = get_column_letter(ArchivoDocentesAdministrativos.Vinculacion + 1)
            nombreVinculacion = str(information[columnNombreVinculacion + str(row)].value)
            
            if "DOCENTE" in nombreVinculacion:
                self.fillDictDocentes(row, dict_Of_Docentes, information)
            else: 
                self.fillDictAdministrativos(row, dict_Of_Administrativos, information)

        print("INCIO GENERACION DE DOCENTES")
        self.generateExcelDocentes(dict_Of_Docentes)
        print("INCIO GENERACION DE ADMINISTRATIVOS")
        self.generateExcelAdministrativos(dict_Of_Administrativos)
    
    def fillDictDocentes(self, row, dict_Of_Docentes, information):
            
            ArchivoDocentesAdministrativos = ArchivosExcel.Docentes

            columnSede = get_column_letter(ArchivoDocentesAdministrativos.Sede + 1)
            sede = str(information[columnSede + str(row)].value)
            
            if sede not in dict_Of_Docentes:
                dict_Of_Docentes[sede] = {}
            
            dict_Of_Facultades = dict_Of_Docentes[sede]
            
            columFacultad = get_column_letter(ArchivoDocentesAdministrativos.Facultad + 1)
            facultad = str(information[columFacultad + str(row)].value)

            if facultad not in dict_Of_Facultades:
                dict_Of_Facultades[facultad] = {}
            
            columUnidad = get_column_letter(ArchivoDocentesAdministrativos.Unidad + 1)
            unidad = str(information[columUnidad + str(row)].value)

            dict_Of_Unidades = dict_Of_Facultades[facultad]

            if unidad not in dict_Of_Unidades:
                dict_Of_Unidades[unidad] = []

            columnCorreo = get_column_letter(ArchivoDocentesAdministrativos.Correo + 1)
            correo = str(information[columnCorreo + str(row)].value)

            dict_Of_Unidades[unidad].append(correo)

    def fillDictAdministrativos(self, row, dict_Of_Administrativos, information):
        
        ArchivoDocentesAdministrativos = ArchivosExcel.Docentes

        columnSede = get_column_letter(ArchivoDocentesAdministrativos.Sede + 1)
        sede = str(information[columnSede + str(row)].value)
            
        if sede not in dict_Of_Administrativos:
            dict_Of_Administrativos[sede] = []

        columnCorreo = get_column_letter(ArchivoDocentesAdministrativos.Correo + 1)
        correo = str(information[columnCorreo + str(row)].value)
        
        dict_Of_Administrativos[sede].append(correo)

    def generateExcelDocentes(self, dict_Of_Docentes):
        for sede in dict_Of_Docentes:
            
            if sede == "SEDE":
                continue
            
            
            print("Rellenar excel " + sede)

            woorkbookSEDE = Workbook()
            woorkbookUNIDAD = Workbook()
            
            hojaSede = woorkbookSEDE.create_sheet(sede)

            dict_sede = dict_Of_Docentes[sede]

            # INSETAR SEDE AL INCIO PARA REUSAR FUNCIONES DE ESTUDIANTES
            sede = "SEDE " + sede.strip()

            usuariosSede = list(dict_sede.keys())
    
            self.fillListaCorreos(hojaSede, sede, usuariosSede, "SEDE", "FACULTAD", sede, "DOCENTE")
            
            for facultad in dict_sede:
                hojaFacultad = woorkbookSEDE.create_sheet(facultad)
                dict_facultad = dict_sede[facultad]
                
                usuariosFacultad = list(dict_facultad.keys())
                self.fillListaCorreos(hojaFacultad, facultad, usuariosFacultad, "FACULTAD", "UNIDAD", sede, "DOCENTE")

                for plan in dict_facultad:
                    hojaPlan = woorkbookUNIDAD.create_sheet(plan)
                    usuariosEstudiantes = dict_facultad[plan]
                    self.fillListaCorreos(hojaPlan, plan, usuariosEstudiantes, "UNIDAD", "DOCENTE", sede,"DOCENTE", facultad)
            
            # NOTACION PARA GUARDAR ARCHIVOS EN WINDOWS ( EN LINUX CAMBIAR )            
            path =  self.folder_pathDocentes + "\\" + sede

            if not os.path.exists(path):
                os.makedirs(path)
            
            woorkbookSEDE.save(path + "\\" + sede + ".xlsx")
            woorkbookUNIDAD.save(path + "\\" + "UNIDADES " + sede + ".xlsx")
    
    def generateExcelAdministrativos(self, dict_Of_Administrativos):
        for sede in dict_Of_Administrativos:
            if sede == "SEDE":
                continue

            woorkbookSEDE = Workbook()
            hojaSede = woorkbookSEDE.create_sheet(sede)

            usuariosSede = dict_Of_Administrativos[sede]

            # INSETAR SEDE AL INCIO PARA REUSAR FUNCIONES DE ESTUDIANTES
            sede = "SEDE " + sede.strip()
    
            self.fillListaCorreos(hojaSede, sede, usuariosSede, "SEDE", "ADMINISTRATIVO", sede, "ADMINISTRATIVO")

            path =  self.folder_pathAdministrativos + "\\" + sede

            if not os.path.exists(path):
                os.makedirs(path)
            
            woorkbookSEDE.save(path + "\\" + sede + ".xlsx")
            
    def fillListaCorreos(self, hoja, GroupMember, users, tipoGroup, tipoUser, sede, tipoArchivo, facultad = None):
        hoja["A1"] = "Group Email"
        hoja["B1"] = "Member Email"
        hoja["C1"] = "Member Type"
        hoja["D1"] = "Member Role"
        hoja["G1"] = "Member NAME"

        row = 2
        userGroupMember = self.get_EmailMember(GroupMember, tipoGroup, sede, tipoArchivo)
        row = self.PropietariosAllListas(hoja, row, userGroupMember)
        row = self.PropietariosSede(hoja, row, userGroupMember, sede)
        
        if tipoGroup == "FACULTAD" or tipoGroup == "UNIDAD":
            row = self.PropietariosFacultad(hoja,userGroupMember, GroupMember, tipoGroup, sede, row, facultad)

        for user in users: 
            hoja["A" + str(row)] = userGroupMember
            hoja["B" + str(row)] = self.get_EmailMember(user, tipoUser, sede, tipoArchivo)
            hoja["C" + str(row)] = "USER" 
            hoja["D" + str(row)] = "MEMBER"
            hoja["G" + str(row)] = user
            row += 1 
            
    def get_EmailMember(self, user : str, tipoUser : str, sede, tipoArchivo : str):
        if tipoUser == "DOCENTE" or tipoUser == "ADMINISTRATIVO":
            return user
        
        if tipoUser == "SEDE":
            # "SEDE BOGOTA"
            sede = user.split(" ")
            # "[SEDE, BOGOTA]"
            sede = sede[1][:3].lower()
            # "bog"
            return tipoArchivo.lower() + "_" + sede + "@unal.edu.co"

         # "SEDE BOGOTA"
        sede = sede.split(" ")
        # "[SEDE, BOGOTA]"
        sede = sede[1][:3].lower()
        # "bog"  

        user = user.split(" ")
        acronimo = ""

        if tipoUser == "FACULTAD":
        
            if (sede == "ama" or sede == "car" 
                or sede == "ori" or sede == "tum"):
                return "estf_" + sede + "@unal.edu.co"

            for palabra in user:
                if len(palabra) > 2:
                    acronimo += palabra.lower()[0]    
            
            return "estf" + acronimo + "_" + sede + "@unal.edu.co"
        
        if tipoUser == "UNIDAD":
            for palabra in user:
                if len(palabra) > 2:
                    acronimo += palabra.capitalize()[:3]
            
            return acronimo + "_" + sede + "@unal.edu.co"
    
    def PropietariosAllListas(self, hoja, row, userGroupMember):
    
        listaNacional = [
        "boletin_un@unal.edu.co",
        "comdninfoa_nal@unal.edu.co",
        "enviosvri_nal@unal.edu.co",
        "rectorinforma@unal.edu.co",
        "comunicado_csu_bog@unal.edu.co",
        "reconsejobu_nal@unal.edu.co",
        "dninfoacad_nal@unal.edu.co",
        "dgt_dned@unal.edu.co",
        "gruposeguridad_nal@unal.edu.co",
        "sisii_nal@unal.edu.co",
        "postmaster_unal@unal.edu.co",
        "postmasterdnia_nal@unal.edu.co",
        "protecdatos_na@unal.edu.co",
        "infraestructurati_dned@unal.edu.co",
        "dre@unal.edu.co",
        "dned@unal.edu.co",

        # Representacion estudiantil 
        "estudiantilcsu@unal.edu.co",
        "estudiantilca@unal.edu.co"
        ]
        
        for owner in listaNacional:
            hoja["A" + str(row)] = userGroupMember
            hoja["B" + str(row)] = owner
            hoja["C" + str(row)] = "USER" 
            hoja["D" + str(row)] = "OWNER"
            hoja["G" + str(row)] = "OWNER COLOMBIA"
            row += 1 
        
        return row
    
    def PropietariosSede(self, hoja, row, userGroupMember, sede):
        lista_sede = []
        
        if sede == self.medellin:
            lista_sede = [
                "alertas_med@unal.edu.co",
                "informa_biblioteca@unal.edu.co",
                "informa_comunicaciones@unal.edu.co",
                "informa_direccion_administrativa@unal.edu.co",
                "informa_direccion_laboratorios@unal.edu.co",
                "informa_fac_ciencias_humanas_y_economicas@unal.edu.co",
                "informa_juridica@unal.edu.co",
                "inf_aplicaciones_med@unal.edu.co",
                "informa_vicerrectoria@unal.edu.co",
                "informa_bienestar_universitario@unal.edu.co",
                "infservcomp_med@unal.edu.co",
                "inflogistica_med@unal.edu.co",
                "informa_fac_ciencias@unal.edu.co",
                "informa_fac_minas@unal.edu.co",
                "informa_fac_ciencias_agrarias@unal.edu.co",
                'info_aplica_med@unal.edu.co',
                "informa_secretaria_sede@unal.edu.co",
                "innovaacad_med@unal.edu.co",
                "unalternativac_nal@unal.edu.co",
                "pcm@unal.edu.co",
                "postmaster_med@unal.edu.co",
                "infeducontinua@unal.edu.co",
                "informa_direccion_academica@unal.edu.co",
                "informa_direccion_de_investigacion_y_extension@unal.edu.co",
                "informa_direccion_ordenamiento_y_desarrollo_fisico@unal.edu.co",
                "informa_fac_arquitectura@unal.edu.co",
                "informa_registro_y_matricula@unal.edu.co",
                "informa_unimedios@unal.edu.co",
                "infpersonal_med@unal.edu.co",
                # Repesentacion estudiantil
            ]

        if sede == self.manizales:
            lista_sede = [
                "ventanilla_man@unal.edu.co",
                "bienestar_man@unal.edu.co",
                "planea_man@unal.edu.co",
                "postmaster_man@unal.edu.co",
                "vicsede_man@unal.edu.co",
                "personaladm_man@unal.edu.co",
                "personaldoc_man@unal.edu.co",
                "saludocup_man@unal.edu.co",
                # Repesentacion manizales
            
            ]
        
        if sede == self.palmira:
            lista_sede = [
                "unnoticias_pal@unal.edu.co",
                "postmaster_pal@unal.edu.co",
                # Representacion 
                
            ]

        if sede == self.orinoquia:
            lista_sede = [
                "divcultural_ori@unal.edu.co"
            ]

        if sede == self.paz:
            lista_sede = [
                "secsedelapaz@unal.edu.co",
                "sedelapaz@unal.edu.co",
                "tics_paz@unal.edu.co",
                "vicesedelapaz@unal.edu.co"
            ]

        if sede == self.bogota:
            lista_sede = [
                "divulgaciondrm_bog@unal.edu.co",
                "talenhumano_bog@unal.edu.co",
                "reprecarrera_bog@unal.edu.co",
                "comunicaciones_bog@unal.edu.co",
                "diracasede_bog@unal.edu.co",
                "dircultural_bog@unal.edu.co",
                "notificass_bog@unal.edu.co",
                "personaladm_bog@unal.edu.co",
                "postmaster_bog@unal.edu.co",
                "salarialp_bog@unal.edu.co"
            ]
        
        for owner in lista_sede:
            hoja["A" + str(row)] = userGroupMember
            hoja["B" + str(row)] = owner
            hoja["C" + str(row)] = "USER" 
            hoja["D" + str(row)] = "OWNER"
            hoja["G" + str(row)] = "OWNER SEDE"
            row += 1 
        
        return row

    def PropietariosFacultad(self, hoja, userGroupMember, GroupMember, tipoGroup, sede, row, facultad):
        
        if sede != self.bogota:
            return row
        
        if tipoGroup == "UNIDAD":
            facultad = facultad
        
        if tipoGroup == "FACULTAD":
            facultad = GroupMember

        
        FacultadBogota = {
            "FACULTAD DE CIENCIAS HUMANAS" : "correo_fchbog@unal.edu.co",
            "FACULTAD DE INGENIERÍA" : "correo_fibog@unal.edu.co",
            "FACULTAD DE INGENIERIA" : "correo_fibog@unal.edu.co",
            "FACULTAD DE CIENCIAS" : "correo_fcbog@unal.edu.co",
            "FACULTAD DE ARTES" : "correo_farbog@unal.edu.co",
            "FACULTAD DE CIENCIAS ECONÓMICAS" : "correo_fcebog@unal.edu.co",
            "FACULTAD DE MEDICINA" : "correo_fmbog@unal.edu.co ",
            "FACULTAD DE DERECHO, CIENCIAS POLÍTICAS Y SOCIALES" : "correo_fdbog@unal.edu.co",
            "FACULTAD DE DERECHO, CIENCIAS POLITICAS Y SOCIALES" : "correo_fdbog@unal.edu.co",
            "FACULTAD DE MEDICINA VETERINARIA Y DE ZOOTECNIA" : "correo_fmvbog@unal.edu.co",
            "FACULTAD DE MEDICINA VETERINARIA Y ZOOTECNICA" : "correo_fmvbog@unal.edu.co",
            "FACULTAD DE CIENCIAS AGRARIAS" : "correo_fcabog@unal.edu.co",
            "FACULTAD DE ENFERMERÍA" : "correo_febog@unal.edu.co",
            "FACULTAD DE ENFERMERIA" : "correo_febog@unal.edu.co",
            "FACULTAD DE ODONTOLOGÍA" : "correo_fobog@unal.edu.co",
            "FACULTAD DE ODONTOLOGIA" : "correo_fobog@unal.edu.co",
            "INSTITUTO DE BIOTECNOLOGÍA - IBUN" : "",
            "INSTITUTO DE ESTUDIOS POLÍTICOS Y RELACIONES INTERNACIONALES - IEPRI" : "",
            "INSTITUTO DE ESTUDIOS URBANOS - IEU" : "",
            "FACULTAD DE CIENCIAS ECONOMICAS" : "",
            "INSTITUTO DE CIENCIA Y TECNOLOGÍA DE ALIMENTOS - ICTA" : "",
            "INSTITUTO DE ESTUDIOS AMBIENTALES - IDEA" : "",
            "INSTITUTO DE GENÉTICA" : "",
            "INSTITUTO DE ESTUDIOS EN COMUNICACIÓN Y CULTURA - IECO" : ""

        }

        hoja["A" + str(row)] = userGroupMember
        hoja["B" + str(row)] = FacultadBogota[facultad]
        hoja["C" + str(row)] = "USER" 
        hoja["D" + str(row)] = "OWNER"
        hoja["G" + str(row)] = "OWNER SEDE"
        row += 1

        return row 
        

    # ------------- Manejo de excel ----------------------------------

    def print_data(self, nombreHoja):
        
        '''
        Se va a recorrer la hoja de exel extrayendo la cantidad de filas 
        atravez de la libreria, mietras que la cantidad de columnas la 
        extraemos dependiendo de hoja que se esta recorriendo. 
        '''

        information = self.excel[nombreHoja]
        cantOfRows = len(list(information.rows))
        cantOfColumns = getCantOfColumns(nombreHoja)
        maxColumns = self.columnaInicial + cantOfColumns
        filaInicial = self.filaInicial

        Datos = False
        for row in range(filaInicial ,cantOfRows):
            for column in range(self.columnaInicial, maxColumns):
                columnChar = get_column_letter(column)
                value = information[columnChar + str(row)].value
                print("cell "+columnChar + str(row), end= " : ")
                print(str(value) + " | ", end= " ")
            print()
            print("--------")
            print()

        print("Cantidad de filas : ")
        print(cantOfRows)
        print("Cantidad de columnas : ")
        print(maxColumns)

