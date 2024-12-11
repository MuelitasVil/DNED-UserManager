from app.services.excel.utils import getCantOfColumns
from app.services.excel.excel.archvivosExcel import ArchivosExcel
from openpyxl.utils import get_column_letter 
from openpyxl import Workbook, load_workbook
import csv
import openpyxl
import os

'''
La siguiente clase tendra la funcionalidad de recorrer el excel en el cual se
encuentran los datos de los inquilinos.

Ojo : Es necesario que entienda la estructura del excel para entender las
siguientes funciones.
'''

class CorreosEstudiantes:
    def __init__(self, file = None, files = []):
        # Lectura del excel 
        if file != None:
            self.excel = openpyxl.load_workbook(file)
            print("HOJAS DEL EXCEL INGRESADO")
            print(self.excel.sheetnames)
            self.hojas = self.excel.get_sheet_names()

        self.folder_path_estudiantes = "archivos\\estudiantes"
        if not os.path.exists(self.folder_path_estudiantes):
            os.makedirs(self.folder_path_estudiantes)

        self.UserFiles = files
        self.columnaInicial = 1
        self.filaInicial = 1   

        self.bogota = "SEDE BOGOTÁ"
        self.amazona = "SEDE AMAZONÍA"
        self.caribe = "SEDE CARIBE"
        self.paz = "SEDE DE LA PAZ" # 
        self.manizales = "SEDE MANIZALES" #
        self.medellin = "SEDE MEDELLÍN" # 
        self.orinoquia = "SEDE ORINOQUÍA" #
        self.palmira = "SEDE PALMIRA" #
        self.tumaco = "SEDE TUMACO"

    def FilterEstudiantes(self):
        
        '''
        Nota : Esta solucion se podria hacer con un arbol binario.
        Sin embargo lo implemente con diccionarios por facilidad

        Estructura : 
        Sedes -> facultades -> planes -> estudiantes

        Dict:
        dict = {
            "sede" : {
                "Facultdad" : {
                    "Plan1" : ["est1","est2","est3"],
                    "Plan1" : ["est1","est2","est3"]
                    }
                }
            }
        '''

        # Datos : 

        ArchivoEstudiantes = ArchivosExcel.EstudiantesActivos 
        
        # OJO : HOJAS DE EXCEL HARDCODEADAS
        information = self.excel["Estudiantes activos"]
        cantOfRows = len(list(information.rows))
        filaInicial = self.filaInicial
 
        #woorkbookOfplans.create_sheet("Informacion General")
        
        # Obtener informacion :
        
        dict_Of_Sedes = {} 
        dict_Of_Sedes_pregrado = {}
        dict_Of_Sedes_postgrado = {}
        
        print("OBTENIENDO INFORMACION : ")
        for row in range(filaInicial ,cantOfRows):

            columnCorreo = get_column_letter(ArchivoEstudiantes.Correo + 1)
            correo = str(information[columnCorreo + str(row)].value)

            if not correo.endswith("@unal.edu.co"):
                continue

            self.fillDictEstudiantes(row, dict_Of_Sedes, information)
            
            columnNivel = get_column_letter(ArchivoEstudiantes.Nivel + 1)
            nivel = str(information[columnNivel + str(row)].value) 
            
            if nivel == "PREGRADO":
                self.fillDictEstudiantes(row, dict_Of_Sedes_pregrado, information)
            else:
                self.fillDictEstudiantes(row, dict_Of_Sedes_postgrado, information)

        print("GENERACION DE EXCEL ESTUDIANTES PREGRADO PROTOTIPO")
        self.generateExcelEstudiantes(dict_Of_Sedes_postgrado, dict_Of_Sedes_pregrado)

    def fillDictEstudiantes(self, row, dict_Of_Sedes, information):
            ArchivoEstudiantes = ArchivosExcel.EstudiantesActivos 

            columnSede = get_column_letter(ArchivoEstudiantes.Sede + 1)
            sede = str(information[columnSede + str(row)].value)
            
            if sede not in dict_Of_Sedes:
                dict_Of_Sedes[sede] = {}
            
            dict_Of_Facultades = dict_Of_Sedes[sede]
            
            columFacultad = get_column_letter(ArchivoEstudiantes.Facultad + 1)
            facultad = str(information[columFacultad + str(row)].value)

            index = facultad.find("(")
            if index != -1:
                facultad = facultad[:index-1]

            if facultad not in dict_Of_Facultades:
                dict_Of_Facultades[facultad] = {}

            dict_planes = dict_Of_Facultades[facultad]
            
            columnPlanEstudio = get_column_letter(ArchivoEstudiantes.CodigoPlan + 1)
            planEstudio = str(information[columnPlanEstudio + str(row)].value)

            columnNombrePlanEstudio = get_column_letter(ArchivoEstudiantes.Plan + 1)
            nombrePlanEstudio = str(information[columnNombrePlanEstudio + str(row)].value)

            if planEstudio not in dict_planes:
                dict_planes[planEstudio] = {}
                dict_planes[planEstudio]['ESTUDIANTES'] = []
                dict_planes[planEstudio]['NOMBRE'] = nombrePlanEstudio

            columnCorreo = get_column_letter(ArchivoEstudiantes.Correo + 1)
            correo = str(information[columnCorreo + str(row)].value)

            dict_planes[planEstudio]['ESTUDIANTES'].append(correo)
        
    def generateExcelEstudiantes(self, dict_Of_Sedes_postgrado : dict, dict_Of_Sedes_pregrado : dict):
        print("\nSEDES EN EL ARCHIVO POSGRADO: ")
        print(list(dict_Of_Sedes_postgrado.keys()))

        print("\nSEDES EN EL ARCHIVO PREGRADO : ")
        print(list(dict_Of_Sedes_pregrado.keys()))

        # Rellenar los excel
        print("\n Rellenar exceles ")

        pregrado = "PREGRADO"
        posgrado = "POSGRADO"

        for sede in dict_Of_Sedes_pregrado:
            if sede == "SEDE":
                continue
            
            print("Rellenar excel " + sede)
            woorkbookFacultad = Workbook()
            woorkbookSEDE = Workbook()
            woorkbookPLAN = Workbook()

            hojaSede = woorkbookSEDE.create_sheet(f"{sede}")

            dict_sede_pregado = dict_Of_Sedes_pregrado[sede]
            facultades = list(dict_sede_pregado.keys())
            
            # Rellenar la informacion de las facultades y los planes informacion pregrado
            self.fillHojasExcel(dict_sede_pregado, woorkbookFacultad, woorkbookPLAN, sede, tipo=pregrado)
            
            usuariosSede_posgrado = []
            if sede in dict_Of_Sedes_postgrado:
                dict_sede_posgrado = dict_Of_Sedes_postgrado[sede]
                usuariosSede_posgrado = list(dict_sede_posgrado.keys())
                # Rellenar la informacion de las facultades y los planes informacion postgrado
                self.fillHojasExcel(dict_sede_posgrado, woorkbookFacultad, woorkbookPLAN, sede, tipo=posgrado)

            # Rellenar la informacion facultades de la sede
            self.fillListaCorreos(hojaSede, sede, facultades, "SEDE", "FACULTAD", sede)
            self.fillHojasFacultades(woorkbookSEDE, sede , facultades)
            # NOTACION PARA GUARDAR ARCHIVOS EN WINDOWS ( EN LINUX CAMBIAR ) 
            path = self.folder_path_estudiantes

            path = path + "\\" + sede
            if not os.path.exists(path):
                os.makedirs(path)
            
            woorkbookSEDE.save(path + "\\" + sede + ".xlsx")
            woorkbookFacultad.save(path + "\\" + "Facultades " + sede + ".xlsx")
            woorkbookPLAN.save(path + "\\" + "PLANES " + sede + ".xlsx")
    
    def fillHojasFacultades(self, woorkbookSEDE : Workbook, sede : str, facultades : list):
        
        for facultad in facultades:
            hojaFacultad = woorkbookSEDE.create_sheet(facultad) 
            facultades = []
            facultadPregrado = self.get_EmailMember(facultad, "FACULTAD", sede, tipoEstudiante="PREGRADO")
            facultadPosgrado =self.get_EmailMember(facultad, "FACULTAD", sede, tipoEstudiante="POSGRADO")

            facultades = [facultadPregrado, facultadPosgrado]
            nuevaHoja = hojaFacultad = woorkbookSEDE.create_sheet(f"{facultad}")

            # Observar que aunque se va a llenar las facultades me aprovecho de que al obtener
            # el usuario estudiantes no lo modifica, por lo tanto puedo pasar directamente los
            # corros de las facultades
            self.fillListaCorreos(nuevaHoja, facultad, facultades, "FACULTAD", "ESTUDIANTE", sede)

    def fillHojasExcel(self, dict_sede, woorkbookFacultad, woorkbookPLAN, sede, tipo):
        for facultad in dict_sede:
                hojaFacultad = woorkbookFacultad.create_sheet(f"{facultad} {tipo}")
                dict_facultad = dict_sede[facultad]
                usuariosFacultad = list(dict_facultad.keys())
                
                #print(f"\nFacultade de {tipo}")
                #print(list(dict_sede.keys()))

                self.fillListaCorreos(hojaFacultad, facultad, usuariosFacultad, "FACULTAD", "PLAN", sede, tipoEstudiante = tipo)
            
                for plan in dict_facultad:
                    infoPlan = dict_facultad[plan]
                    nombrePlan_pregado = dict_facultad[plan]['NOMBRE']
                    
                    hojaPlan_pregado = woorkbookPLAN.create_sheet(plan + " | " + nombrePlan_pregado)

                    usuariosEstudiantes_pregado = infoPlan['ESTUDIANTES']
                    self.fillListaCorreos(hojaPlan_pregado, plan, usuariosEstudiantes_pregado, "PLAN", "ESTUDIANTE", sede, facultad, tipoEstudiante = tipo)

    def fillListaCorreos(self, hoja, GroupMember, users, tipoGroup, tipoUser, sede, facultad = None, row = None, plan = None, tipoEstudiante = None):
        
        datosCsv = []

        hoja["A1"] = "Group Email"
        hoja["B1"] = "Member Email"
        hoja["C1"] = "Member Type"
        hoja["D1"] = "Member Role"
        hoja["G1"] = "Member NAME"
        hoja["H1"] = "PLAN NAME"

        datosCsv.append(["Group Email", "Member Email", "Member Type", "Member Role"])

        userGroupMember = self.get_EmailMember(GroupMember, tipoGroup, sede, tipoEstudiante=tipoEstudiante)
        
        if row == None:
            row = 2
            row = self.PropietariosAllListas(hoja, row, userGroupMember, datosCsv)
            row = self.PropietariosSede(hoja, row, userGroupMember, sede, datosCsv)
        
            if tipoGroup == "FACULTAD" or tipoGroup == "UNIDAD" or tipoGroup == "PLAN":
                row = self.PropietariosFacultad(hoja,userGroupMember, GroupMember, tipoGroup, sede, row, facultad, datosCsv)
        
        for user in users: 
            hoja["A" + str(row)] = userGroupMember
            hoja["B" + str(row)] = self.get_EmailMember(user, tipoUser, sede, tipoEstudiante=tipoEstudiante)
            hoja["C" + str(row)] = "USER" 
            hoja["D" + str(row)] = "MEMBER"

            hoja["G" + str(row)] = user
            hoja["H" + str(row)] = plan
            row += 1 

            datosCsv.append(
                [
                userGroupMember,
                self.get_EmailMember(user, tipoUser, sede, tipoEstudiante=tipoEstudiante),
                "USER", 
                "MEMBER"]
                )

        self.generarCsv(sede, tipoGroup, GroupMember, datosCsv, tipoEstudiante)
        return row
            
    def get_EmailMember(self, user : str, tipoUser : str, sede, tipoEstudiante = None):
        
        if tipoEstudiante == "PREGRADO":
            tipoEstudiante = "pre"
        elif tipoEstudiante == "POSGRADO":
            tipoEstudiante = "pos"
        else:
            tipoEstudiante = ""

        if tipoUser == "ESTUDIANTE":
            return user
        
        if tipoUser == "SEDE":
            # "SEDE BOGOTA"
            sede = user.split(" ")
            # "[SEDE, BOGOTA]"
            sede = sede[1][:3].lower()
            # "bog"
            return f"estudiante{tipoEstudiante}_" + sede  + "@unal.edu.co"

         # "SEDE BOGOTA"
        sede = sede.split(" ")
        # "[SEDE, BOGOTA]"
        sede = sede[1][:3].lower()
        # "bog"  

        if tipoUser == "FACULTAD":
        
            if (sede == "ama" or sede == "car" 
                or sede == "ori" or sede == "tum"):
                return f"estf{tipoEstudiante}" + sede + "@unal.edu.co"

            acronimo = ""
            for palabra in user.split():
                if len(palabra) > 2:
                    acronimo += palabra.lower()[0]    
            
            return f"est{acronimo}{tipoEstudiante}"  + "_" + sede + "@unal.edu.co"
        
        if tipoUser == "PLAN":
             # for palabra in user:
             #   if len(palabra) > 2:
             #       acronimo += palabra.capitalize()[:3]
            
            return user + "_" + sede + "@unal.edu.co"
    
    def PropietariosAllListas(self, hoja, row : int, userGroupMember : str, datosCsv : list):
    
        listaNacional = [
        "boletin_un@unal.edu.co",
        "comdninfoa_nal@unal.edu.co",
        "enviosvri_nal@unal.edu.co",
        "rectorinforma@unal.edu.co",
        "comunicado_csu_bog@unal.edu.co",
        "reconsejobu_nal@unal.edu.co",
        "dninfoacad_nal@unal.edu.co",
        "dgt_dned@unal.edu.co",
        "gruposeguridad_nal@unal.edu.co",
        "sisii_nal@unal.edu.co",
        "postmaster_unal@unal.edu.co",
        "postmasterdnia_nal@unal.edu.co",
        "protecdatos_na@unal.edu.co",
        "infraestructurati_dned@unal.edu.co",
        "dre@unal.edu.co",
        "dned@unal.edu.co",

        # Representacion estudiantil 
        "estudiantilcsu@unal.edu.co",
        "estudiantilca@unal.edu.co"
        ]
        
        for owner in listaNacional:
            hoja["A" + str(row)] = userGroupMember
            hoja["B" + str(row)] = owner
            hoja["C" + str(row)] = "USER" 
            hoja["D" + str(row)] = "OWNER" #OWNER
            hoja["G" + str(row)] = "OWNER COLOMBIA"
            row += 1 

            datosCsv.append(
                [
                userGroupMember,
                owner,
                "USER", 
                "OWNER"]
                )
        
        return row
    
    def PropietariosSede(self, hoja, row : int, userGroupMember : str, sede : list, datosCsv : list):
        lista_sede = []
        
        if sede == self.medellin:
            lista_sede = [
                "alertas_med@unal.edu.co",
                "informa_biblioteca@unal.edu.co",
                "informa_comunicaciones@unal.edu.co",
                "informa_direccion_administrativa@unal.edu.co",
                "informa_direccion_laboratorios@unal.edu.co",
                "informa_fac_ciencias_humanas_y_economicas@unal.edu.co",
                "informa_juridica@unal.edu.co",
                "inf_aplicaciones_med@unal.edu.co",
                "informa_vicerrectoria@unal.edu.co",
                "informa_bienestar_universitario@unal.edu.co",
                "infservcomp_med@unal.edu.co",
                "inflogistica_med@unal.edu.co",
                "informa_fac_ciencias@unal.edu.co",
                "informa_fac_minas@unal.edu.co",
                "informa_fac_ciencias_agrarias@unal.edu.co",
                'info_aplica_med@unal.edu.co',
                "informa_secretaria_sede@unal.edu.co",
                "innovaacad_med@unal.edu.co",
                "unalternativac_nal@unal.edu.co",
                "pcm@unal.edu.co",
                "postmaster_med@unal.edu.co",
                "infeducontinua@unal.edu.co",
                "informa_direccion_academica@unal.edu.co",
                "informa_direccion_de_investigacion_y_extension@unal.edu.co",
                "informa_direccion_ordenamiento_y_desarrollo_fisico@unal.edu.co",
                "informa_fac_arquitectura@unal.edu.co",
                "informa_registro_y_matricula@unal.edu.co",
                "informa_unimedios@unal.edu.co",
                "infpersonal_med@unal.edu.co",
                # Repesentacion estudiantil
                "reestudia_med@unal.edu.co"
            ]

        if sede == self.manizales:
            lista_sede = [
                "ventanilla_man@unal.edu.co",
                "bienestar_man@unal.edu.co",
                "planea_man@unal.edu.co",
                "postmaster_man@unal.edu.co",
                "vicsede_man@unal.edu.co",
                # "personaladm_man@unal.edu.co", # Administrativos
                # "personaldoc_man@unal.edu.co", # Docentes
                # "saludocup_man@unal.edu.co", # Administrativos
                # Repesentacion manizales
                "estudiantilcs_man@unal.edu.co"
            ]
        
        if sede == self.palmira:
            lista_sede = [
                "unnoticias_pal@unal.edu.co",
                "postmaster_pal@unal.edu.co",
                # Representacion 
                "estudiantilcs_pal@unal.edu.co"
            ]

        if sede == self.orinoquia:
            lista_sede = [
                "divcultural_ori@unal.edu.co"
            ]

        if sede == self.paz:
            lista_sede = [
                "secsedelapaz@unal.edu.co",
                "sedelapaz@unal.edu.co",
                "tics_paz@unal.edu.co",
                "vicesedelapaz@unal.edu.co"
            ]

        if sede == self.bogota:
            lista_sede = [
                "divulgaciondrm_bog@unal.edu.co",
                # "talenhumano_bog@unal.edu.co", # Adminsitrativos y docentes
                "reprecarrera_bog@unal.edu.co",
                "comunicaciones_bog@unal.edu.co",
                "diracasede_bog@unal.edu.co",
                "dircultural_bog@unal.edu.co",
                "notificass_bog@unal.edu.co",
                # "personaladm_bog@unal.edu.co", # Administrativos 
                "postmaster_bog@unal.edu.co",
                # "salarialp_bog@unal.edu.co" # Administrativos y docentes 
            ]
        
        for owner in lista_sede:
            hoja["A" + str(row)] = userGroupMember
            hoja["B" + str(row)] = owner
            hoja["C" + str(row)] = "USER" 
            hoja["D" + str(row)] = "OWNER" #OWNER
            hoja["G" + str(row)] = "OWNER SEDE"

            datosCsv.append(
                [
                userGroupMember,
                owner,
                "USER", 
                "OWNER"]
                )

            row += 1 
        
        return row

    def PropietariosFacultad(self, hoja, userGroupMember : str, GroupMember : str, tipoGroup : str, sede : str, row : int, facultad : str, datosCsv : list):
        
        if sede != self.bogota:
            return row
        
        if tipoGroup == "PLAN":
            facultad = facultad
        
        if tipoGroup == "FACULTAD":
            facultad = GroupMember

        FacultadBogota = {
            "FACULTAD DE CIENCIAS HUMANAS" : "correo_fchbog@unal.edu.co",
            "FACULTAD DE INGENIERÍA" : "correo_fibog@unal.edu.co",
            "FACULTAD DE CIENCIAS" : "correo_fcbog@unal.edu.co",
            "FACULTAD DE ARTES" : "correo_farbog@unal.edu.co",
            "FACULTAD DE CIENCIAS ECONÓMICAS" : "correo_fcebog@unal.edu.co",
            "FACULTAD DE MEDICINA" : "correo_fmbog@unal.edu.co ",
            "FACULTAD DE DERECHO, CIENCIAS POLÍTICAS Y SOCIALES" : "correo_fdbog@unal.edu.co",
            "FACULTAD DE MEDICINA VETERINARIA Y DE ZOOTECNIA" : "correo_fmvbog@unal.edu.co",
            "FACULTAD DE CIENCIAS AGRARIAS" : "correo_fcabog@unal.edu.co",
            "FACULTAD DE ENFERMERÍA" : "correo_febog@unal.edu.co",
            "FACULTAD DE ODONTOLOGÍA" : "correo_fobog@unal.edu.co",
            "FACULTAD DE ODONTOLOGÍA" : "correo_fobog@unal.edu.co",
            
            # Preguntar al inge
            "FACULTAD DE CIENCIAS HUMANAS (Admisión PAET)" : "correo_fchbog@unal.edu.co",
            "FACULTAD DE INGENIERÍA (Admisión PAET)" : "correo_fibog@unal.edu.co",
            "FACULTAD DE CIENCIAS (Admisión PAET)" : "correo_fcbog@unal.edu.co",
            "FACULTAD DE ARTES (Admisión PAET)" : "correo_farbog@unal.edu.co",
            "FACULTAD DE CIENCIAS ECONÓMICAS (Admisión PAET)" : "correo_fcebog@unal.edu.co",
            "FACULTAD DE MEDICINA (Admisión PAET)" : "correo_fmbog@unal.edu.co ",
            "FACULTAD DE DERECHO, CIENCIAS POLÍTICAS Y SOCIALES (Admisión PAET)" : "correo_fdbog@unal.edu.co",
            "FACULTAD DE MEDICINA VETERINARIA Y DE ZOOTECNIA (Admisión PAET)" : "correo_fmvbog@unal.edu.co",
            "FACULTAD DE CIENCIAS AGRARIAS (Admisión PAET)" : "correo_fcabog@unal.edu.co",
            "FACULTAD DE ENFERMERÍA (Admisión PAET)" : "correo_febog@unal.edu.co",
            "FACULTAD DE ODONTOLOGÍA (Admisión PAET)" : "correo_fobog@unal.edu.co",
            "FACULTAD DE ODONTOLOGÍA (Admisión PAET)" : "correo_fobog@unal.edu.co",
        }

        hoja["A" + str(row)] = userGroupMember
        hoja["B" + str(row)] = FacultadBogota[facultad]
        hoja["C" + str(row)] = "USER" 
        hoja["D" + str(row)] = "OWNER" #OWNER
        hoja["G" + str(row)] = "OWNER SEDE"
        row += 1

        datosCsv.append(
                [
                userGroupMember,
                FacultadBogota[facultad],
                "USER", 
                "OWNER"]
                )

        return row 
        

    # ------------- Manejo de excel ----------------------------------

    def print_data(self, nombreHoja):
        
        '''
        Se va a recorrer la hoja de exel extrayendo la cantidad de filas 
        atravez de la libreria, mietras que la cantidad de columnas la 
        extraemos dependiendo de hoja que se esta recorriendo. 
        '''

        information = self.excel[nombreHoja]
        cantOfRows = len(list(information.rows))
        cantOfColumns = getCantOfColumns(nombreHoja)
        maxColumns = self.columnaInicial + cantOfColumns
        filaInicial = self.filaInicial

        Datos = False
        for row in range(filaInicial ,cantOfRows):
            for column in range(self.columnaInicial, maxColumns):
                columnChar = get_column_letter(column)
                value = information[columnChar + str(row)].value
                print("cell "+columnChar + str(row), end= " : ")
                print(str(value) + " | ", end= " ")
            print()
            print("--------")
            print()

        print("Cantidad de filas : ")
        print(cantOfRows)
        print("Cantidad de columnas : ")
        print(maxColumns)


#Example write in csv

    def generarCsv(self,sede, tipoGroup, GroupMember, datosCsv, tipoEstudiante): 
    # Abre los archivos CSV en modo de escritura
        path = self.folder_path_estudiantes
        path = f"{path}\\{sede}\\{tipoGroup}-csv\\"
        if tipoGroup == "FACULTAD" and tipoEstudiante != None:
            path = path = path + f"{tipoEstudiante}\\"
        
        if not os.path.exists(path):
                os.makedirs(path)
        
        path = path + GroupMember + ".csv"
        with open(path, mode='w', newline='') as file1:
            writerUsers = csv.writer(file1)
            # Itera sobre las filas de datos
            for dato in datosCsv:
                # Escribe la fila en ambos archivos CSV
                writerUsers.writerow(dato)
    

'''
    def generateExcelEstudiantesFacultad(self, dict_Of_Sedes, tipo):
        print("SEDES EN EL ARCHIVO : ")
        print(list(dict_Of_Sedes.keys()))

        print("FACULTADES BOGOTA : ")
        print(list(dict_Of_Sedes["SEDE BOGOTÁ"].keys()))

        # Rellenar los excel
        print("Rellenar exceles ")
        for sede in dict_Of_Sedes:
            
            if sede == "SEDE":
                continue
            
            print("Rellenar excel " + sede)
            woorkbookSEDE = Workbook()
            hojaSede = woorkbookSEDE.create_sheet(sede)

            dict_sede = dict_Of_Sedes[sede]
            usuariosSede = list(dict_sede.keys())
            
            self.fillListaCorreos(hojaSede, sede, usuariosSede, "SEDE", "FACULTAD", sede)
            
            for facultad in dict_sede:
                hojaFacultad = woorkbookSEDE.create_sheet(facultad)
                dict_facultad = dict_sede[facultad]
                
                usuariosFacultad = list(dict_facultad.keys())
                row = self.fillListaCorreos(hojaFacultad, facultad, usuariosFacultad, "FACULTAD", "PLAN", sede)
                
                for plan in dict_facultad:
                    infoPlan = dict_facultad[plan]
                    nombrePlan = dict_facultad[plan]['NOMBRE']
                    usuariosEstudiantes = infoPlan['ESTUDIANTES']
                    userPlan = str(plan) + " | " +nombrePlan

                    row = self.fillListaCorreos(hojaFacultad, facultad, usuariosEstudiantes, "FACULTAD", "ESTUDIANTE", sede, facultad, row=row, plan = userPlan)

            # NOTACION PARA GUARDAR ARCHIVOS EN WINDOWS ( EN LINUX CAMBIAR )            
            if tipo == "POSGRADO":
                path =  self.folder_path_estudiantes_postgrado

            path = path + "\\" + sede
            if not os.path.exists(path):
                os.makedirs(path)
            
            woorkbookSEDE.save(path + "\\" + sede + ".xlsx")
    
    '''
    