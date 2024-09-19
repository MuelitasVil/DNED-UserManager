from app.services.excel.utils import getCantOfColumns
from app.services.excel.excel.archvivosExcel import ArchivosExcel
from io import StringIO
import time
from openpyxl.utils import get_column_letter
import openpyxl 
from openpyxl import Workbook, load_workbook

import csv

'''
La siguiente clase tendra la funcionalidad de recorrer el excel en el cual se
encuentran los datos de los inquilinos.

Ojo : Es necesario que entienda la estructura del excel para entender las
siguientes funciones.
'''

class MergeUsers:
    def __init__(self, file = None, files = []):
        # Lectura del excel 
        if file != None:
            self.excel = openpyxl.load_workbook(file)
            self.hojas = self.excel.get_sheet_names()

        self.UserFiles = files
        self.columnaInicial = 1
        self.filaInicial = 1    
    
    def MergueUsuarios(self):
        
        '''
        Diccionario = {
            CorreoUsuario : {
                Nombre : String
                Espacio Gastado : 
            }
        }
        '''

        inicio = time.time()
        print(" Inicio del mergue ")
        Usuarios = {}
        for file in self.UserFiles:
            print(file)
            if file.filename.endswith(".xlsx") or file.filename.endswith(".xlsm") :
                excelFile = openpyxl.load_workbook(file)
                self.getDataUsersExel(excelFile, file.filename, Usuarios)
            else:
                # Es csv LDAP
                self.getDataUsersCsv(file, Usuarios)

        print(" Estructura de datos generada ")
        print(" Inicio generar excel ")
        
        # Limberar espacio ram
        excelFile = None

        # Generar el excel con la informacion 
        ws = Workbook()
        hojaUsuario = ws.create_sheet("Informacion General")
        hojaEgresado = ws.create_sheet("Egresados")
        hojaNOEgresado = ws.create_sheet("NO Egresados")
        hojaLdap = ws.create_sheet("OnlyLdap")
        hojaWorkSpace = ws.create_sheet("OnlyWorkSpace")
        hojaEstudiante = ws.create_sheet("Estudiante")
        hojaPregrado = ws.create_sheet("Pregrado")
        hojaPostGrado = ws.create_sheet("Postgrado")
        hojaDocente = ws.create_sheet("Docente")
        hojaPensionado = ws.create_sheet("Pensionado")
        hojaAministrativo = ws.create_sheet("Administrativo")
        hojaContratista = ws.create_sheet("Contratista")
        hojaSinConexion = ws.create_sheet("No conexion")
        
        # Generar el excel con mas de 100G
        wsGrande = Workbook()
        hojaUsuarioG = wsGrande.create_sheet("Informacion General")
        hojaEgresadoG = wsGrande.create_sheet("Only Egresados")
        hojaNOEgresadoG = wsGrande.create_sheet("NO Egresados")
        hojaLdapG = wsGrande.create_sheet("Only Ldap")
        hojaWorkSpaceG = wsGrande.create_sheet("Only WorkSpace")
        hojaEstudianteG = wsGrande.create_sheet("Estudiante")
        hojaPregradoG = wsGrande.create_sheet("Pregrado")
        hojaPostGradoG = wsGrande.create_sheet("Postgrado")
        hojaDocenteG = wsGrande.create_sheet("Docente")
        hojaPensionadoG = wsGrande.create_sheet("Pensionado")
        hojaAministrativoG = wsGrande.create_sheet("Administrativo")
        hojaContratistaG = wsGrande.create_sheet("Contratista")
        hojaSinConexionG = wsGrande.create_sheet("No conexion")

        columnas = self.get_columnas_usuarios()
        cantidadColumnas = len(columnas)

        # Rellenar columnas de los archivos
        for c in range(cantidadColumnas):
            column = get_column_letter(c + 1)
            hojaUsuario[column + "1"].value = columnas[c]
            hojaUsuarioG[column + "1"].value = columnas[c]
            if c <= 4:
                hojaEgresado[column + "1"].value = columnas[c]
                hojaNOEgresado[column + "1"].value = columnas[c]
                hojaWorkSpace[column + "1"].value = columnas[c]
                hojaLdap[column + "1"].value = columnas[c]
                hojaEstudiante[column + "1"].value = columnas[c]
                hojaPregrado[column + "1"].value = columnas[c]
                hojaPostGrado[column + "1"].value = columnas[c]
                hojaDocente[column + "1"].value = columnas[c]
                hojaPensionado[column + "1"].value = columnas[c]
                hojaAministrativo[column + "1"].value = columnas[c]
                hojaContratista[column + "1"].value = columnas[c]
                hojaSinConexion[column + "1"].value = columnas[c]

                hojaEgresadoG[column + "1"].value = columnas[c]
                hojaNOEgresadoG[column + "1"].value = columnas[c]
                hojaWorkSpaceG[column + "1"].value = columnas[c]
                hojaLdapG[column + "1"].value = columnas[c]
                hojaEstudianteG[column + "1"].value = columnas[c]
                hojaPregradoG[column + "1"].value = columnas[c]
                hojaPostGradoG[column + "1"].value = columnas[c]
                hojaDocenteG[column + "1"].value = columnas[c]
                hojaPensionadoG[column + "1"].value = columnas[c]
                hojaAministrativoG[column + "1"].value = columnas[c]
                hojaContratistaG[column + "1"].value = columnas[c]
                hojaSinConexionG[column + "1"].value = columnas[c]

        # Ingresar Columnas 
        rowUsuario, rowUsuarioG = 2, 2
        rowEgresado, rowEgresadoG = 2, 2
        rowNOEgresado, rowNOEgresadoG = 2, 2 
        rowLdap, rowLdapG = 2, 2 
        rowWorkSpace, rowWorkSpaceG = 2, 2 
        rowEstudiante, rowEstudianteG = 2, 2
        rowPregrado, rowPregradoG = 2, 2
        rowPostgrado, rowPostgradoG = 2, 2 
        rowDocente, rowDocenteG  = 2, 2 
        rowPensionado, rowPensionadoG = 2, 2    
        rowAdministrativo, rowAdministrativoG = 2, 2
        rowContratista, rowContratistaG = 2, 2
        rowNoConexion, rowNoConexionG = 2, 2

        for Correo in Usuarios:

            if rowUsuario % 50000 == 0:
                print("Generando ... ")

            rowUsuario, rowUsuarioG = self.FillUser(hojaUsuario, hojaUsuarioG, Usuarios, rowUsuario, rowUsuarioG, Correo)
            
            if Usuarios[Correo]["OnlyWorkSpace"]:
                rowWorkSpace, rowWorkSpaceG = self.FillTypeUser(hojaWorkSpace, hojaWorkSpaceG, Usuarios, rowWorkSpace,rowWorkSpaceG, Correo)
            elif Usuarios[Correo]["OnlyLdap"]:
                rowLdap, rowLdapG = self.FillTypeUser(hojaLdap, hojaLdapG,Usuarios, rowLdap, rowLdapG, Correo)
            elif Usuarios[Correo]["IsEgresado"] :
                rowEgresado, rowEgresadoG = self.FillTypeUser(hojaEgresado, hojaEgresadoG,Usuarios, rowEgresado, rowEgresadoG, Correo)
            
            if not Usuarios[Correo]["IsEgresado"] or Usuarios[Correo]["OnlyLdap"] or Usuarios[Correo]["OnlyWorkSpace"]:
                rowNOEgresado, rowNOEgresadoG = self.FillTypeUser(hojaNOEgresado, hojaNOEgresadoG,Usuarios, rowNOEgresado, rowNOEgresadoG, Correo)

            if Usuarios[Correo]["IsEstudiante"]:
                rowEstudiante, rowEstudianteG = self.FillTypeUser(hojaEstudiante, hojaEstudianteG, Usuarios, rowEstudiante, rowEstudianteG, Correo)
                
            if Usuarios[Correo]["IsPregrado"]:
                rowPregrado, rowPregradoG = self.FillTypeUser(hojaPregrado, hojaPregradoG, Usuarios, rowPregrado, rowPregradoG, Correo)
            
            if Usuarios[Correo]["IsPostgrado"]:
                rowPostgrado, rowPostgradoG = self.FillTypeUser(hojaPostGrado, hojaPostGradoG, Usuarios, rowPostgrado, rowPostgradoG, Correo)
            
            if Usuarios[Correo]["IsDocente"]: 
                rowDocente, rowDocenteG = self.FillTypeUser(hojaDocente, hojaDocenteG, Usuarios, rowDocente, rowDocenteG, Correo)
            
            if Usuarios[Correo]["IsPensionado"]:
                rowPensionado, rowPensionadoG = self.FillTypeUser(hojaPensionado, hojaPensionadoG, Usuarios, rowPensionado, rowPensionadoG, Correo)

            if Usuarios[Correo]["IsAdministrativo"]:
                rowAdministrativo, rowAdministrativoG = self.FillTypeUser(hojaAministrativo, hojaAministrativoG, Usuarios, rowAdministrativo, rowAdministrativoG, Correo)
            
            if Usuarios[Correo]["IsContratista"]:
                rowContratista, rowContratistaG = self.FillTypeUser(hojaContratista, hojaContratistaG, Usuarios, rowContratista, rowContratistaG, Correo)
            
            if Usuarios[Correo]["NoConexion"]:
                rowNoConexion, rowNoConexionG = self.FillTypeUser(hojaSinConexion, hojaSinConexionG, Usuarios, rowNoConexion, rowNoConexionG, Correo)

        print(" Excel generado ")
        print(" Guardando Excel ")
        
        ws.save("Usuarios.xlsx")
        wsGrande.save("Usuarios_100G.xlsx")
        fin = time.time()   
        
        print("Fin")
        print("Tiempo tomado : " + str(round(fin-inicio)))
        
        return False
    
    def FillUser(self, hoja, hoja100G, Usuarios, row, rowG, Correo):
        row = row + 1 
        hoja["A" + str(row)] = Correo
        hoja["B" + str(row)] = Usuarios[Correo]["Nombre"]
        hoja["C" + str(row)] = Usuarios[Correo]["Apellido"] 
        hoja["D" + str(row)] = Usuarios[Correo]["UltimaConexion"] 
        hoja["E" + str(row)] = Usuarios[Correo]["Almacenamiento"]
        hoja["F" + str(row)] = Usuarios[Correo][ArchivosExcel.Docentes.TipoArchivo]
        hoja["G" + str(row)] = Usuarios[Correo][ArchivosExcel.Egresados.TipoArchivo] 
        hoja["H" + str(row)] = Usuarios[Correo][ArchivosExcel.EstudiantesActivos.TipoArchivo]
        hoja["I" + str(row)] = Usuarios[Correo][ArchivosExcel.WorkSpace.TipoArchivo] 
        hoja["J" + str(row)] = Usuarios[Correo][ArchivosExcel.Ldap.TipoArchivo] 
        hoja["K" + str(row)] = Usuarios[Correo]["IsEgresado"] 
        
        try:
            if float(Usuarios[Correo]["Almacenamiento"]) >= 100:
                rowG = rowG + 1 
                hoja100G["A" + str(rowG)] = Correo
                hoja100G["B" + str(rowG)] = Usuarios[Correo]["Nombre"]
                hoja100G["C" + str(rowG)] = Usuarios[Correo]["Apellido"] 
                hoja100G["D" + str(rowG)] = Usuarios[Correo]["UltimaConexion"] 
                hoja100G["E" + str(rowG)] = Usuarios[Correo]["Almacenamiento"]
                hoja100G["F" + str(rowG)] = Usuarios[Correo][ArchivosExcel.Docentes.TipoArchivo]
                hoja100G["G" + str(rowG)] = Usuarios[Correo][ArchivosExcel.Egresados.TipoArchivo] 
                hoja100G["H" + str(rowG)] = Usuarios[Correo][ArchivosExcel.EstudiantesActivos.TipoArchivo]
                hoja100G["I" + str(rowG)] = Usuarios[Correo][ArchivosExcel.WorkSpace.TipoArchivo] 
                hoja100G["J" + str(rowG)] = Usuarios[Correo][ArchivosExcel.Ldap.TipoArchivo] 
                hoja100G["K" + str(rowG)] = Usuarios[Correo]["IsEgresado"] 
                return (row, rowG)
        except ValueError:
            return (row, rowG)
    
        return (row, rowG)


    def FillTypeUser(self, hoja, hoja100G, Usuarios, row, rowG, Correo):
        row = row + 1 

        hoja["A" + str(row)] = Correo
        hoja["B" + str(row)] = Usuarios[Correo]["Nombre"]
        hoja["C" + str(row)] = Usuarios[Correo]["Apellido"] 
        hoja["D" + str(row)] = Usuarios[Correo]["UltimaConexion"] 
        hoja["E" + str(row)] = Usuarios[Correo]["Almacenamiento"]
        hoja["F" + str(row)] = Usuarios[Correo][ArchivosExcel.Ldap.TipoArchivo]
            
        try:
            if float(Usuarios[Correo]["Almacenamiento"]) >= 100:
                rowG = rowG + 1
                hoja100G["A" + str(rowG)] = Correo
                hoja100G["B" + str(rowG)] = Usuarios[Correo]["Nombre"]
                hoja100G["C" + str(rowG)] = Usuarios[Correo]["Apellido"] 
                hoja100G["D" + str(rowG)] = Usuarios[Correo]["UltimaConexion"] 
                hoja100G["E" + str(rowG)] = Usuarios[Correo]["Almacenamiento"]
                hoja100G["F" + str(rowG)] = Usuarios[Correo][ArchivosExcel.Ldap.TipoArchivo]
                return (row, rowG)
        except ValueError:
            return (row, rowG)
    
        return (row, rowG)

    def getDataUsersCsv(self, csvFile, Usuarios):
        # Leer el contenido del archivo y decodificarlo a cadena de texto
        csv_content = csvFile.read().decode('utf-8')

        # Usar StringIO para crear un archivo en memoria a partir de la cadena de texto
        csv_file = StringIO(csv_content)

        # Crear el lector CSV
        csvReader = csv.reader(csv_file)

        ldapAchivo = ArchivosExcel.Ldap
        tipos = ArchivosExcel.Ldap.TipoUsuario
        for row in csvReader:
            Correo = row[ldapAchivo.Correo]
            gropu_tipo = row[ldapAchivo.Tipo]
            
            #if not Correo.endswith("@unal.edu.co"):
            #    continue
            
            if Correo not in Usuarios:
                self.create_user_dict(Correo, Usuarios)

            TiposUsuario = ""
            array_tipos = gropu_tipo.split("|")
            for tipo in array_tipos: 

                if str(tipo) not in tipos:
                    continue

                tipoUsuario = tipos[str(tipo)]

                if tipoUsuario == "Estudiante":
                    Usuarios[Correo]["IsEstudiante"] = True

                if tipoUsuario == "Docente":
                    Usuarios[Correo]["IsDocente"] = True

                if tipoUsuario == "Administrativo":
                    Usuarios[Correo]["IsAdministrativo"] = True

                if tipoUsuario == "Contratista":
                    Usuarios[Correo]["IsContratista"] = True

                if tipoUsuario == "Pensionado": 
                    Usuarios[Correo]["IsPensionado"] = True

                if tipoUsuario != "Egresado":
                    Usuarios[Correo]["IsEgresado"] = False
                    
                Usuarios[Correo]["OnlyWorkSpace"] = False

                TiposUsuario = TiposUsuario + " | " + tipoUsuario

            Usuarios[Correo]["Ldap"] = TiposUsuario

    def getDataUsersExel(self, excelFile, filename, Usuarios):
        # Encontrar cual Archivo es 
        nombreHoja = ""
        if filename == ArchivosExcel.Docentes.NombreArchivo:
            Archivo = ArchivosExcel.Docentes
            nombreHoja = Archivo.NombreHoja
        if filename == ArchivosExcel.Egresados.NombreArchivo:
            Archivo = ArchivosExcel.Egresados
            nombreHoja = Archivo.NombreHoja
        if filename == ArchivosExcel.EstudiantesActivos.NombreArchivo:
            Archivo = ArchivosExcel.EstudiantesActivos
            nombreHoja = Archivo.NombreHoja
        if filename == ArchivosExcel.WorkSpace.NombreArchivo:
            Archivo = ArchivosExcel.WorkSpace
            nombreHoja = Archivo.NombreHoja

        information = excelFile.get_sheet_by_name(nombreHoja)
        cantOfRows = len(list(information.rows))
        filaInicial = self.filaInicial

        for row in range(filaInicial ,cantOfRows):
            columnCorreo = get_column_letter(Archivo.Correo + 1)
            Correo = information[columnCorreo + str(row)].value
            
            if Correo not in Usuarios:
                self.create_user_dict(Correo, Usuarios)
                
            if ArchivosExcel.WorkSpace.TipoArchivo == Archivo.TipoArchivo:
                columnNombre = get_column_letter(Archivo.Nombre + 1)
                columnApellido = get_column_letter(Archivo.Apellidos + 1)
                columnUltimaConexion = get_column_letter(Archivo.UltimaConexion + 1)
                columnAlmacenamiento = get_column_letter(Archivo.Almacenamiento + 1)
                
                Nombre = information[columnNombre + str(row)].value
                Apellido = information[columnApellido + str(row)].value
                UltimaConexion = information[columnUltimaConexion + str(row)].value
                Almacenamiento = information[columnAlmacenamiento + str(row)].value

                Usuarios[Correo]["Nombre"] = Nombre
                Usuarios[Correo]["Apellido"] = Apellido
                Usuarios[Correo]["UltimaConexion"] = UltimaConexion
                Usuarios[Correo]["Almacenamiento"] = Almacenamiento

            if (Archivo.TipoArchivo != ArchivosExcel.WorkSpace.TipoArchivo): 
                Usuarios[Correo]["OnlyWorkSpace"] = False

            if (Archivo.TipoArchivo != ArchivosExcel.Ldap.TipoArchivo):
                Usuarios[Correo]["OnlyLdap"] = False

            if (Archivo.TipoArchivo == ArchivosExcel.EstudiantesActivos.TipoArchivo):
                Usuarios[Correo]["IsEstudiante"] = True

            if (Archivo.TipoArchivo != ArchivosExcel.Egresados.TipoArchivo
                and Archivo.TipoArchivo != ArchivosExcel.WorkSpace.TipoArchivo):
                 Usuarios[Correo]["IsEgresado"] = False

            if Archivo.TipoArchivo == ArchivosExcel.EstudiantesActivos.TipoArchivo:
                columnNivel = get_column_letter(Archivo.Nivel + 1)
                Nivel = str(information[columnNivel + str(row)].value).strip()
                if "PREGRADO" in Nivel:
                    Usuarios[Correo]["IsPregrado"] = True
                elif "POSGRADO" in Nivel:
                    Usuarios[Correo]["IsPostgrado"] = True

            if Archivo.TipoArchivo == ArchivosExcel.Docentes.TipoArchivo:
                columnVinculacion = get_column_letter(Archivo.Vinculacion + 1)
                Vinculacion = str(information[columnVinculacion + str(row)].value)
                if "DOCENTE" in Vinculacion:
                    Usuarios[Correo]["IsDocente"] = True
                else:
                    Usuarios[Correo]["IsAdministrativo"] = True

            if Archivo.TipoArchivo == ArchivosExcel.WorkSpace.TipoArchivo:
                columnConexion = get_column_letter(Archivo.UltimaConexion + 1)
                Conexion = str(information[columnConexion + str(row)].value)
                if Conexion.strip() == "Never logged in":
                    Usuarios[Correo]["NoConexion"] = True

            Usuarios[Correo][Archivo.TipoArchivo] = True
        
    def create_user_dict(self, Correo, Usuarios):
        Usuarios[Correo] = {}
        Usuarios[Correo][ArchivosExcel.Docentes.TipoArchivo] = False
        Usuarios[Correo][ArchivosExcel.Egresados.TipoArchivo] = False
        Usuarios[Correo][ArchivosExcel.EstudiantesActivos.TipoArchivo] = False
        Usuarios[Correo][ArchivosExcel.WorkSpace.TipoArchivo] = False
        Usuarios[Correo][ArchivosExcel.Ldap.TipoArchivo] = False
        Usuarios[Correo]["Ldap"] = "" 
        Usuarios[Correo]["Nombre"] = ""
        Usuarios[Correo]["Apellido"] = ""
        Usuarios[Correo]["UltimaConexion"] = ""
        Usuarios[Correo]["Almacenamiento"] = ""
        Usuarios[Correo]["IsEgresado"] = True
        Usuarios[Correo]["IsEstudiante"] = False
        Usuarios[Correo]["IsPregrado"] = False
        Usuarios[Correo]["IsPostgrado"] = False
        Usuarios[Correo]["IsDocente"] = False
        Usuarios[Correo]["IsPensionado"] = False
        Usuarios[Correo]["IsContratista"] = False
        Usuarios[Correo]["IsAdministrativo"] = False
        Usuarios[Correo]["OnlyWorkSpace"] = True
        Usuarios[Correo]["OnlyLdap"] = True
        Usuarios[Correo]["NoConexion"] = False

    def get_columnas_usuarios(self):

        Columnas = [
            "Correo Usuario",
            "Nombre",
            "Apellido",
            "UltimaConexion",
            "Almacenamiento",
            ArchivosExcel.Docentes.TipoArchivo,
            ArchivosExcel.Egresados.TipoArchivo,
            ArchivosExcel.EstudiantesActivos.TipoArchivo,
            ArchivosExcel.WorkSpace.TipoArchivo,
            ArchivosExcel.Ldap.TipoArchivo,
            "NO CONEXION"
            ]
        
        return Columnas
            
    # ------------- Manejo de excel ----------------------------------

    def print_data(self, nombreHoja):
        
        '''
        Se va a recorrer la hoja de exel extrayendo la cantidad de filas 
        atravez de la libreria, mietras que la cantidad de columnas la 
        extraemos dependiendo de hoja que se esta recorriendo. 
        '''

        information = self.excel.get_sheet_by_name(nombreHoja)
        cantOfRows = len(list(information.rows))
        cantOfColumns = getCantOfColumns(nombreHoja)
        maxColumns = self.columnaInicial + cantOfColumns
        filaInicial = self.filaInicial

        Datos = False
        for row in range(filaInicial ,cantOfRows):
            for column in range(self.columnaInicial, maxColumns):
                columnChar = get_column_letter(column)
                value = information[columnChar + str(row)].value
                print("cell "+columnChar + str(row), end= " : ")
                print(str(value) + " | ", end= " ")
            print()
            print("--------")
            print()

        print("Cantidad de filas : ")
        print(cantOfRows)
        print("Cantidad de columnas : ")
        print(maxColumns)


'''
Example write in csv

 # Abre los archivos CSV en modo de escritura
        with open(file_users, mode='w', newline='') as file1, open(file_csv, mode='w', newline='') as file2:
            writerUsers = csv.writer(file1)
            writerCsv = csv.writer(file2)
    
            # Escribe los encabezados de las columnas en ambos archivos
            writerUsers.writerow(columns)
            writerCsv.writerow(columsCsv)

            # Itera sobre las filas de datos
            for row in range(filaInicial, cantOfRows):
                indexEmail = get_column_letter(2)
                email = information[indexEmail + str(row)].value

                indexSede = get_column_letter(3)
                sede = information[indexSede + str(row)].value

                indexFacultad = get_column_letter(4)
                facultad = information[indexFacultad + str(row)].value

                indexCod_plan = get_column_letter(5)
                cod_plan = information[indexCod_plan + str(row)].value

                indexPlan = get_column_letter(6)
                plan = information[indexPlan + str(row)].value

                indexNivel = get_column_letter(6)
                nivel = information[indexNivel + str(row)].value

                # Filtros

                if sede not in ["SEDE BOGOTÁ"]:
                    continue
                
                if facultad not in ['FACULTAD DE INGENIERÍA']:
                    continue
                
                # Escribe la fila en ambos archivos CSV
                row_user = [email, sede, facultad, cod_plan, plan, nivel]
                row_csv = list(infoCsv.values())
                row_csv[1] = email 

                writerUsers.writerow(row_user)
                writerCsv.writerow(row_csv)
    
'''